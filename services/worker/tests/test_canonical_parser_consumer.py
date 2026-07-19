from pathlib import Path

import pytest
from openpyxl import Workbook

from app.canonical.md_sanitizer import sanitize_markdown
from app.canonical.text_to_md import text_to_markdown
from app.canonical.to_markdown import canonicalize
from app.canonical.xlsx_to_md import xlsx_to_markdown
from app.consumer import process_ingest_job
from app.models import DocumentNode, TextChunk
from app.parser.document_parser import clean_text, parse_document, parse_excel, parse_text
from app.profile_index_plan import ProfileIndexPlan


def test_sanitize_and_text_to_markdown(tmp_path):
    p = tmp_path / "a.txt"
    p.write_text("#Title  \r\n\r\n\r\nbody  \n", encoding="utf-8")

    assert sanitize_markdown("") == ""
    assert sanitize_markdown("#Title  \r\n\r\n\r\nbody  \n") == "# Title\n\nbody"
    assert text_to_markdown(p) == "# Title\n\nbody"


def test_xlsx_to_markdown_escapes_tables_and_skips_empty_sheets(tmp_path):
    path = tmp_path / "book.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["A|B", "C"])
    ws.append(["x\ny", 2])
    wb.create_sheet("Empty")
    wb.save(path)

    md = xlsx_to_markdown(path)

    assert "## Data" in md
    assert "A\\|B" in md
    assert "x y" in md
    assert "Empty" not in md


def test_canonicalize_routes_by_suffix_and_mime(tmp_path):
    text = tmp_path / "note.md"
    text.write_text("hello", encoding="utf-8")
    pdf = tmp_path / "doc.bin"
    pdf.write_text("pdf", encoding="utf-8")
    docx = tmp_path / "doc.unknown"
    docx.write_text("docx", encoding="utf-8")
    xlsx = tmp_path / "book.unknown"
    xlsx.write_text("xlsx", encoding="utf-8")

    assert canonicalize(text, "application/octet-stream", "note.md") == "hello"
    import app.canonical.to_markdown as router
    router.pdf_to_markdown = lambda path: f"pdf:{path.name}"
    router.docx_to_markdown = lambda path: f"docx:{path.name}"
    router.xlsx_to_markdown = lambda path: f"xlsx:{path.name}"
    assert canonicalize(pdf, "application/pdf", "doc.bin") == "pdf:doc.bin"
    assert canonicalize(docx, "application/vnd.openxmlformats-officedocument.wordprocessingml.document", "doc.bin") == "docx:doc.unknown"
    assert canonicalize(xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "doc.bin") == "xlsx:book.unknown"
    with pytest.raises(ValueError, match="Unsupported file type"):
        canonicalize(tmp_path / "bad.bin", "application/octet-stream", "bad.bin")


def test_parse_text_excel_and_unsupported_document(tmp_path):
    txt = tmp_path / "a.txt"
    txt.write_text("a\r\n\r\n\r\nb", encoding="utf-8")
    assert clean_text("a\r\n\t b\n\n\nc") == "a\n b\n\nc"
    assert parse_text(txt)[0].text == "a\n\nb"
    assert parse_document(txt, "text/plain", "a.txt")[0].metadata["source"] == "a.txt"

    xlsx = tmp_path / "a.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    ws.append(["A", "B"])
    ws.append([1, None])
    wb.save(xlsx)

    nodes = parse_excel(xlsx)
    assert nodes[0].metadata == {"sheet": "S", "source": "a.xlsx", "type": "excel"}
    assert "A | B" in nodes[0].text
    assert parse_document(xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "a.xlsx")

    with pytest.raises(ValueError):
        parse_document(tmp_path / "a.bin", "application/octet-stream", "a.bin")


def test_process_ingest_job_success_uses_canonical_pipeline(monkeypatch):
    statuses = []
    index_calls = []
    chunks = [TextChunk("c1", 0, "content", 2, {"heading": "H"})]

    monkeypatch.setattr("app.consumer.post_status", lambda payload: statuses.append(payload))
    monkeypatch.setattr("app.consumer.download_object", lambda object_key, dest: Path(dest).write_text("raw", encoding="utf-8"))
    monkeypatch.setattr("app.consumer.canonicalize", lambda path, mime, name: "# Title\nbody")
    monkeypatch.setattr(
        "app.consumer.build_profile_index_plan",
        lambda *args, **kwargs: ProfileIndexPlan(chunks, chunks, chunks),
    )

    class FakeEmbedder:
        def __init__(self, model):
            self.model = model

        def embed(self, text):
            return [1.0, 2.0]

        def embed_batch(self, texts):
            return [[1.0, 2.0]]

    class FakeIndexer:
        def __init__(self, dimension, collection_name=None, profile_schema=False):
            self.dimension = dimension
            self.profile_schema = profile_schema

        def delete_by_doc(self, doc_id):
            index_calls.append(("profile-delete" if self.profile_schema else "legacy-delete", doc_id))

        def index_chunks(self, kb_id, doc_id, chunks_arg, vectors):
            index_calls.append(("legacy-index", [chunk.id for chunk in chunks_arg]))
            chunks_arg[0].milvus_id = "m1"
            return ["m1"]

        def index_profile_chunks(self, kb_id, doc_id, chunks_arg, vectors):
            index_calls.append(("profile-index", [chunk.id for chunk in chunks_arg]))
            chunks_arg[0].milvus_id = "m1"
            return ["m1"]

    monkeypatch.setattr("app.consumer.Embedder", FakeEmbedder)
    monkeypatch.setattr("app.consumer.MilvusIndexer", FakeIndexer)

    process_ingest_job({
        "jobId": "j",
        "kbId": "kb",
        "docId": "doc",
        "objectKey": "obj",
        "fileName": "a.txt",
        "mimeType": "text/plain",
        "embeddingDimension": 2,
    })

    assert [s.get("stage") for s in statuses if "stage" in s] == ["parsing", "chunking", "embedding", "indexing", "completed"]
    assert statuses[-1]["chunks"][0]["milvusId"] == "m1"
    assert statuses[-1]["indexSchemaVersion"] == 2
    assert index_calls == [
        ("profile-delete", "doc"),
        ("profile-index", ["c1"]),
        ("legacy-delete", "doc"),
        ("legacy-index", ["c1"]),
    ]


def test_process_ingest_job_falls_back_and_reports_failures(monkeypatch):
    statuses = []
    monkeypatch.setattr("app.consumer.post_status", lambda payload: statuses.append(payload))
    monkeypatch.setattr("app.consumer.download_object", lambda object_key, dest: Path(dest).write_text("raw", encoding="utf-8"))
    monkeypatch.setattr("app.consumer.canonicalize", lambda *args: (_ for _ in ()).throw(ValueError("unsupported")))
    monkeypatch.setattr("app.consumer.parse_document", lambda *args: [DocumentNode("raw")])
    monkeypatch.setattr(
        "app.consumer.build_profile_index_plan",
        lambda *args, **kwargs: ProfileIndexPlan([], [], []),
    )

    process_ingest_job({
        "jobId": "j",
        "kbId": "kb",
        "docId": "doc",
        "objectKey": "obj",
        "fileName": "a.bin",
        "mimeType": "application/octet-stream",
    })

    assert statuses[-1]["status"] == "failed"
    assert "No chunks produced" in statuses[-1]["errorMessage"]


def test_process_ingest_job_reports_embedding_dimension_mismatch(monkeypatch):
    statuses = []
    chunks = [TextChunk("c1", 0, "content", 2, {})]
    monkeypatch.setattr("app.consumer.post_status", lambda payload: statuses.append(payload))
    monkeypatch.setattr("app.consumer.download_object", lambda object_key, dest: Path(dest).write_text("raw", encoding="utf-8"))
    monkeypatch.setattr("app.consumer.canonicalize", lambda path, mime, name: "body")
    monkeypatch.setattr(
        "app.consumer.build_profile_index_plan",
        lambda *args, **kwargs: ProfileIndexPlan(chunks, chunks, chunks),
    )

    class FakeEmbedder:
        def __init__(self, model):
            self.model = model

        def embed(self, text):
            return [1.0]

        def embed_batch(self, texts):
            return [[1.0]]

    monkeypatch.setattr("app.consumer.Embedder", FakeEmbedder)

    process_ingest_job({
        "jobId": "j",
        "kbId": "kb",
        "docId": "doc",
        "objectKey": "obj",
        "fileName": "a.txt",
        "mimeType": "text/plain",
        "embeddingDimension": 2,
    })

    assert statuses[-1]["status"] == "failed"
    assert "Embedding dimension mismatch" in statuses[-1]["errorMessage"]


def test_process_ingest_job_parent_child_indexes_children_and_reports_parents(monkeypatch):
    statuses = []
    embedded_texts = []
    indexed_chunk_ids = []
    parent = TextChunk("p1", 0, "full parent context", 4, {"chunk_role": "parent"})
    child = TextChunk(
        "c1",
        1,
        "child",
        1,
        {"chunk_role": "child", "parent_chunk_id": "p1"},
    )

    monkeypatch.setattr("app.consumer.post_status", lambda payload: statuses.append(payload))
    monkeypatch.setattr(
        "app.consumer.download_object",
        lambda object_key, dest: Path(dest).write_text("raw", encoding="utf-8"),
    )
    monkeypatch.setattr("app.consumer.canonicalize", lambda path, mime, name: "body")
    monkeypatch.setattr(
        "app.consumer.build_profile_index_plan",
        lambda *args, **kwargs: ProfileIndexPlan(
            [parent, child],
            [child],
            [],
        ),
    )

    class FakeEmbedder:
        def __init__(self, model):
            self.model = model

        def embed(self, text):
            return [1.0, 2.0]

        def embed_batch(self, texts):
            embedded_texts.extend(texts)
            return [[1.0, 2.0] for _ in texts]

    class FakeIndexer:
        def __init__(self, dimension, collection_name=None, profile_schema=False):
            self.dimension = dimension

        def delete_by_doc(self, doc_id):
            pass

        def index_chunks(self, kb_id, doc_id, chunks_arg, vectors):
            pass

        def index_profile_chunks(self, kb_id, doc_id, chunks_arg, vectors):
            indexed_chunk_ids.extend(chunk.id for chunk in chunks_arg)
            for chunk in chunks_arg:
                chunk.milvus_id = f"m-{chunk.id}"
            return [chunk.milvus_id for chunk in chunks_arg]

    monkeypatch.setattr("app.consumer.Embedder", FakeEmbedder)
    monkeypatch.setattr("app.consumer.MilvusIndexer", FakeIndexer)

    process_ingest_job({
        "jobId": "j",
        "kbId": "kb",
        "docId": "doc",
        "objectKey": "obj",
        "fileName": "a.txt",
        "mimeType": "text/plain",
        "embeddingDimension": 2,
        "retrievalProfile": "parent-child",
    })

    assert embedded_texts == ["child"]
    assert indexed_chunk_ids == ["c1"]
    assert [chunk["id"] for chunk in statuses[-1]["chunks"]] == ["p1", "c1"]
    assert statuses[-1]["chunks"][1]["metadata"]["parent_chunk_id"] == "p1"


def test_process_ingest_job_qa_assisted_indexes_original_and_qa_chunks(monkeypatch):
    statuses = []
    embedded_texts = []
    indexed_chunk_ids = []
    source = TextChunk("source-1", 0, "source content", 2, {"source": "a.txt"})
    qa_chunk = TextChunk(
        "qa-1",
        1,
        "Question: Q?\nAnswer: A.",
        4,
        {"chunk_role": "qa", "source_chunk_id": "source-1"},
    )

    monkeypatch.setattr("app.consumer.post_status", lambda payload: statuses.append(payload))
    monkeypatch.setattr("app.consumer.download_object", lambda key, dest: Path(dest).write_text("raw", encoding="utf-8"))
    monkeypatch.setattr("app.consumer.canonicalize", lambda *args: "body")
    monkeypatch.setattr(
        "app.consumer.build_profile_index_plan",
        lambda *args, **kwargs: ProfileIndexPlan(
            [source, qa_chunk],
            [source, qa_chunk],
            [source],
        ),
    )

    class FakeEmbedder:
        def __init__(self, model):
            self.model = model

        def embed(self, text):
            return [1.0, 2.0]

        def embed_batch(self, texts):
            embedded_texts.extend(texts)
            return [[1.0, 2.0] for _ in texts]

    class FakeIndexer:
        def __init__(self, dimension, collection_name=None, profile_schema=False):
            self.dimension = dimension

        def delete_by_doc(self, doc_id):
            pass

        def index_chunks(self, kb_id, doc_id, chunks_arg, vectors):
            pass

        def index_profile_chunks(self, kb_id, doc_id, chunks_arg, vectors):
            indexed_chunk_ids.extend(chunk.id for chunk in chunks_arg)
            for chunk in chunks_arg:
                chunk.milvus_id = f"m-{chunk.id}"

    monkeypatch.setattr("app.consumer.Embedder", FakeEmbedder)
    monkeypatch.setattr("app.consumer.MilvusIndexer", FakeIndexer)

    process_ingest_job({
        "jobId": "j",
        "kbId": "kb",
        "docId": "doc",
        "objectKey": "obj",
        "fileName": "a.txt",
        "mimeType": "text/plain",
        "embeddingDimension": 2,
        "retrievalProfile": "qa-assisted",
    })

    assert embedded_texts == ["source content", "Question: Q?\nAnswer: A."]
    assert indexed_chunk_ids == ["source-1", "qa-1"]
    assert [chunk["id"] for chunk in statuses[-1]["chunks"]] == ["source-1", "qa-1"]
    assert statuses[-1]["status"] == "completed"


def test_process_ingest_job_combined_uses_parents_as_qa_sources(monkeypatch):
    statuses = []
    indexed_chunk_ids = []
    parent = TextChunk("parent-1", 0, "parent context", 3, {"chunk_role": "parent"})
    child = TextChunk("child-1", 1, "child content", 2, {"chunk_role": "child", "parent_chunk_id": "parent-1"})
    qa_chunk = TextChunk("qa-1", 2, "Question: Q?\nAnswer: A.", 4, {"chunk_role": "qa", "source_chunk_id": "parent-1"})

    monkeypatch.setattr("app.consumer.post_status", lambda payload: statuses.append(payload))
    monkeypatch.setattr("app.consumer.download_object", lambda key, dest: Path(dest).write_text("raw", encoding="utf-8"))
    monkeypatch.setattr("app.consumer.canonicalize", lambda *args: "body")
    monkeypatch.setattr(
        "app.consumer.build_profile_index_plan",
        lambda *args, **kwargs: ProfileIndexPlan(
            [parent, child, qa_chunk],
            [child, qa_chunk],
            [],
        ),
    )

    class FakeEmbedder:
        def __init__(self, model):
            self.model = model

        def embed(self, text):
            return [1.0, 2.0]

        def embed_batch(self, texts):
            return [[1.0, 2.0] for _ in texts]

    class FakeIndexer:
        def __init__(self, dimension, collection_name=None, profile_schema=False):
            self.dimension = dimension

        def delete_by_doc(self, doc_id):
            pass

        def index_chunks(self, kb_id, doc_id, chunks_arg, vectors):
            pass

        def index_profile_chunks(self, kb_id, doc_id, chunks_arg, vectors):
            indexed_chunk_ids.extend(chunk.id for chunk in chunks_arg)

    monkeypatch.setattr("app.consumer.Embedder", FakeEmbedder)
    monkeypatch.setattr("app.consumer.MilvusIndexer", FakeIndexer)

    process_ingest_job({
        "jobId": "j",
        "kbId": "kb",
        "docId": "doc",
        "objectKey": "obj",
        "fileName": "a.txt",
        "mimeType": "text/plain",
        "embeddingDimension": 2,
        "retrievalProfile": "combined",
    })

    assert indexed_chunk_ids == ["child-1", "qa-1"]
    assert [chunk["id"] for chunk in statuses[-1]["chunks"]] == ["parent-1", "child-1", "qa-1"]


def test_process_ingest_job_qa_failure_still_completes_original_index(monkeypatch):
    statuses = []
    indexed_chunk_ids = []
    indexer_modes = []
    source = TextChunk("source-1", 0, "source content", 2, {})

    monkeypatch.setattr("app.consumer.post_status", lambda payload: statuses.append(payload))
    monkeypatch.setattr("app.consumer.download_object", lambda key, dest: Path(dest).write_text("raw", encoding="utf-8"))
    monkeypatch.setattr("app.consumer.canonicalize", lambda *args: "body")
    monkeypatch.setattr(
        "app.consumer.build_profile_index_plan",
        lambda *args, **kwargs: ProfileIndexPlan([source], [source], [source]),
    )

    class FakeEmbedder:
        def __init__(self, model):
            self.model = model

        def embed(self, text):
            return [1.0, 2.0]

        def embed_batch(self, texts):
            return [[1.0, 2.0] for _ in texts]

    class FakeIndexer:
        def __init__(self, dimension, collection_name=None, profile_schema=False):
            self.dimension = dimension
            indexer_modes.append(profile_schema)

        def delete_by_doc(self, doc_id):
            pass

        def index_chunks(self, kb_id, doc_id, chunks_arg, vectors):
            pass

        def index_profile_chunks(self, kb_id, doc_id, chunks_arg, vectors):
            indexed_chunk_ids.extend(chunk.id for chunk in chunks_arg)

    monkeypatch.setattr("app.consumer.Embedder", FakeEmbedder)
    monkeypatch.setattr("app.consumer.MilvusIndexer", FakeIndexer)

    process_ingest_job({
        "jobId": "j",
        "kbId": "kb",
        "docId": "doc",
        "objectKey": "obj",
        "fileName": "a.txt",
        "mimeType": "text/plain",
        "embeddingDimension": 2,
        "retrievalProfile": "qa-assisted",
        "legacyWriteRequired": False,
    })

    assert indexed_chunk_ids == ["source-1"]
    assert indexer_modes == [True]
    assert statuses[-1]["status"] == "completed"
    assert [chunk["id"] for chunk in statuses[-1]["chunks"]] == ["source-1"]
