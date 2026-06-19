import re
from pathlib import Path

import fitz
from docx import Document as DocxDocument
from openpyxl import load_workbook

from app.models import DocumentNode


def clean_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def parse_pdf(path: Path) -> list[DocumentNode]:
    nodes: list[DocumentNode] = []
    with fitz.open(path) as doc:
        for page_num, page in enumerate(doc, start=1):
            text = clean_text(page.get_text("text"))
            if text:
                nodes.append(DocumentNode(text=text, metadata={"page": page_num, "source": path.name}))
    return nodes


def parse_docx(path: Path) -> list[DocumentNode]:
    doc = DocxDocument(path)
    nodes: list[DocumentNode] = []
    heading = ""
    for para in doc.paragraphs:
        text = clean_text(para.text)
        if not text:
            continue
        style = (para.style.name if para.style else "").lower()
        if "heading" in style:
            heading = text
            nodes.append(DocumentNode(text=text, metadata={"heading": heading, "source": path.name}))
        else:
            nodes.append(DocumentNode(text=text, metadata={"heading": heading, "source": path.name}))
    return nodes


def parse_text(path: Path) -> list[DocumentNode]:
    text = clean_text(path.read_text(encoding="utf-8", errors="ignore"))
    return [DocumentNode(text=text, metadata={"source": path.name})] if text else []


def parse_excel(path: Path) -> list[DocumentNode]:
    wb = load_workbook(path, read_only=True, data_only=True)
    nodes: list[DocumentNode] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if cells:
                rows.append(" | ".join(cells))
        if rows:
            text = clean_text("\n".join(rows))
            nodes.append(DocumentNode(
                text=text,
                metadata={"sheet": sheet_name, "source": path.name, "type": "excel"},
            ))
    wb.close()
    return nodes


def parse_document(path: Path, mime_type: str, file_name: str) -> list[DocumentNode]:
    suffix = path.suffix.lower()
    if suffix == ".pdf" or "pdf" in mime_type:
        return parse_pdf(path)
    if suffix in (".docx",) or "wordprocessingml" in mime_type:
        return parse_docx(path)
    if suffix in (".xlsx", ".xls") or "spreadsheet" in mime_type:
        return parse_excel(path)
    if suffix in (".txt", ".md", ".markdown") or mime_type.startswith("text/"):
        return parse_text(path)
    raise ValueError(f"Unsupported file type: {file_name} ({mime_type})")
