from pathlib import Path

from openpyxl import load_workbook

from app.canonical.md_sanitizer import sanitize_markdown


def _escape_cell(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", " ")


def xlsx_to_markdown(path: Path) -> str:
    wb = load_workbook(path, read_only=True, data_only=True)
    sections: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = [_escape_cell(str(c).strip()) for c in row if c is not None and str(c).strip()]
            if cells:
                rows.append(cells)

        if not rows:
            continue

        lines = [f"## {sheet_name}", ""]
        header = rows[0]
        lines.append("| " + " | ".join(header) + " |")
        lines.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in rows[1:]:
            padded = row + [""] * (len(header) - len(row))
            lines.append("| " + " | ".join(padded[:len(header)]) + " |")

        sections.append("\n".join(lines))

    wb.close()
    return sanitize_markdown("\n\n".join(sections))
